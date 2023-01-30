from django.contrib import admin

from app_main.models import Category, GeneralData, Product, Banner, Suscriptor, InfoUtil, Municipio, Orden, \
    ComponenteOrden, ContenidoInfo


class CategoryAdmin(admin.ModelAdmin):
    list_display = ('name', 'img_link', 'destacado')
    fields = ('name', 'image', 'destacado')
    actions = ['Activar_destacados', 'Quitar_destacados']

    def Activar_destacados(self, request, queryset):
        for p in queryset:
            p.destacado = True
            p.save()

    def Quitar_destacados(self, request, queryset):
        for p in queryset:
            p.destacado = False
            p.save()


class BannerInline(admin.TabularInline):
    model = Banner
    extra = 5
    fields = ('banner',)


class ContenidoInfoInline(admin.StackedInline):
    model = ContenidoInfo
    extra = 3
    fields = ('text', 'image', 'image_tag')
    readonly_fields = ('image_tag',)


class GeneralDataAdmin(admin.ModelAdmin):
    list_display = (
        'enterprise_name', 'email', 'phone_number', 'taza_cambio', 'tropipay_impuesto', 'logo_link',
        'img_principal_link',
    )
    fieldsets = [
        ('Datos principales', {
            'fields': ('enterprise_name', 'taza_cambio', 'tropipay_impuesto', 'logo', 'img_principal',)
        },),
        ('Redes Sociales', {
            'fields': ('facebook', 'instagram',)
        },),
        ('Contacto', {
            'fields': ('enterprise_address', 'email', 'phone_number')
        },),
    ]
    inlines = [BannerInline]

    def has_add_permission(self, request):
        user = request.user
        if not user.is_anonymous:
            return False if GeneralData.objects.filter().exists() else True
        return False


class ProductInline(admin.StackedInline):
    model = Product
    extra = 1
    fieldsets = [
        ('Datos Principales:', {
            'fields': ('name', 'category', 'specifications', 'price', 'old_price')
        }),
        ('Descripción:', {
            'fields': ('info', 'about')
        })
    ]


class ProductAdmin(admin.ModelAdmin):
    list_display = (
        'name', 'category', 'img_link', 'moneda', 'price', 'info_tag', 'sales', 'stock', 'has_old_price', 'is_active')
    fieldsets = [
        ('Datos Principales:', {
            'fields': ('name', 'category', 'moneda', 'price', 'old_price', 'stock', 'delivery_time')
        }),
        ('Descripción:', {
            'fields': ('image', 'img_link', 'is_active', 'is_important', 'info', 'about')
        })
    ]
    search_fields = ('name',)
    list_filter = ('category',)
    actions = ['Desactivar_productos', 'Activar_productos', 'Activar_destacados', 'Quitar_destacados',
               'Quitar_descuento']
    change_list_template = 'admin/custom_list.html'
    readonly_fields = ('date_updated', 'sales', 'img_link')

    def Desactivar_productos(self, request, queryset):
        for p in queryset:
            p.is_active = False
            p.save()

    def Activar_productos(self, request, queryset):
        for p in queryset:
            p.is_active = True
            p.save()

    def Activar_destacados(self, request, queryset):
        for p in queryset:
            p.is_important = False
            p.save()

    def Quitar_destacados(self, request, queryset):
        for p in queryset:
            p.is_important = True
            p.save()

    def Quitar_descuento(self, request, queryset):
        for p in queryset:
            p.old_price = None
            p.save()


class InfoUtilAdmin(admin.ModelAdmin):
    list_display = ('title',)
    inlines = [ContenidoInfoInline]


class MunicipioAdmin(admin.ModelAdmin):
    list_display = ('nombre', 'precio', 'precio_euro')


class OrdenAdmin(admin.ModelAdmin):
    list_display = ('status', 'uuid', 'date_created', 'total', 'moneda', 'correo', 'municipio')
    list_filter = ('status', 'moneda', 'municipio')
    search_fields = ('uuid',)
    actions = ['Exportar_Excel']

    def Exportar_Excel(self, request, queryset):
        import datetime
        from django.http import HttpResponse
        from openpyxl import Workbook
        from io import BytesIO

        qs = Orden.objects.all()
        filas = [i[0] for i in set(queryset.values_list('correo'))]
        print(filas)
        columnas = [Product.objects.get(pk=i[0]).name for i in
                    set(ComponenteOrden.objects.filter(orden__in=queryset).values_list('producto_id'))]
        filas2 = []
        for i in filas:
            a = Orden.objects.filter(correo=i).first()
            filas2.append(f'{a.nombre_receptor}\n {a.telefono_receptor}\n {a.municipio}')
        print(columnas)

        # book = openpyxl.Workbook()
        # sheet = book.active
        # export = [columnas]
        #
        # sheet.append(columnas)
        #
        # for ob in object_list:
        #     sheet.append((
        #         f'{ob["nombre"]} {ob["apellidos"]}',
        #         ob["categoria_ocupacional"],
        #         ob['categoria_cientifica'],
        #         ob['rrhh']['institucion'],
        #         ob['rrhh']['clasificador_entidad'],
        #         int(ob['rrhh']['porciento_de_participacion']),
        #         ob['rrhh']['salario_mensual'],
        #         ob['rrhh']['salario_anual_ejecutora'],
        #         ob['rrhh']['salario_anual_externo'],
        #         int(ob['rrhh']['porciento_de_remuneracion']),
        #         ob['rrhh']['mce'],
        #         ob['rrhh']['tiempo'],
        #         ob['rrhh']['anual'],
        #         ob['rrhh']['salario_mensual_basico'],
        #     ))
        #
        # today = datetime.now()
        # strToday = today.strftime("%Y-%m-%d")
        #
        # sheet = excel.pe.Sheet(export)
        #
        # return excel.make_response(sheet, "xlsx",
        #                                file_name=f"Resumen de salarios {project} " + strToday + ".xlsx")
        #
        # sheet.title = f'{project}'
        # response = HttpResponse(content_type='application / msexcel')
        # response['Content-Disposition'] = f'attachment;filename=Anexo 3 {project} {strToday}.xlsx'
        # book.save(response)
        # return response
        wb = Workbook()  # Genere un libro de trabajo (es decir, un archivo de Excel)
        wb.encoding = 'utf-8'
        sheet1 = wb.active  # Obtenga la primera hoja de trabajo (hoja1)
        sheet1.title = 'Ordenes'
        sheet1.append([''] + columnas)

        filas1 = [''] + filas2
        for i in range(1, len(filas1) + 1):
            sheet1.cell(row=i, column=1).value = filas1[i - 1]

        output = BytesIO()
        wb.save(output)  # Guarde el contenido del archivo de Excel en IO

        output.seek(0)  # Reposición al principio
        response = HttpResponse(output.getvalue(), content_type='application/vnd.ms-excel')
        ctime = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        file_name = 'Información registrada % s.xls' % ctime
        response['Content-Disposition'] = 'attachment; filename=%s' % file_name
        return response
        for i in qs:
            for a in i.componente_orden.all():
                pass
        print(ComponenteOrden.objects.filter(orden__in=Orden.objects.all()).count())
        # for i in ComponenteOrden.objects.filter(orden__in=queryset):
        #     print(i)


class ComponenteOrdenAdmin(admin.ModelAdmin):
    list_display = ('orden', 'Componente')
    list_filter = ('producto',)
    search_fields = ('orden', 'producto')


admin.site.register(Product, ProductAdmin)
admin.site.register(InfoUtil, InfoUtilAdmin)
admin.site.register(Category, CategoryAdmin)
admin.site.register(GeneralData, GeneralDataAdmin)
admin.site.register(Suscriptor)
admin.site.register(Orden, OrdenAdmin)
admin.site.register(ComponenteOrden, ComponenteOrdenAdmin)
admin.site.register(Municipio, MunicipioAdmin)
